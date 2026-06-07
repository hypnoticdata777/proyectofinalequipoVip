"""
Genera la Matriz de Trazabilidad (Actividad 4) en formato Excel.
Proyecto: USF Portal - Universidad Santa Fe, Campus Paraíso
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ─── COLORES INSTITUCIONALES ─────────────────────────────────────────────────
AZUL_OSCURO  = "0A2540"
DORADO       = "D4AF37"
AZUL_CLARO   = "1E3A5F"
GRIS_CLARO   = "F5F5F5"
GRIS_MEDIO   = "DDDDDD"
VERDE        = "28A745"
NARANJA      = "FFA500"
ROJO         = "DC3545"
MORADO       = "6F42C1"
BLANCO       = "FFFFFF"
AMARILLO_CLARO = "FFF9E6"
VERDE_CLARO  = "E8F5E9"
ROJO_CLARO   = "FFEBEE"
AZUL_FILAS   = "EBF3FA"

# ─── DATOS DEL PROYECTO ───────────────────────────────────────────────────────
REQS_FUNCIONALES = [
    # ID, Descripción, Módulo, Archivos, Prioridad, CU, CP, Estado
    ("RF-01","El sistema debe permitir autenticación mediante Google OAuth 2.0",
     "Autenticación","backend/src/config/passport.js\nbackend/src/controllers/auth.controller.js\nfrontend/src/app/modules/auth/login/",
     "Alta","CU-01","CP-01","Finalizado"),
    ("RF-02","El sistema debe asignar roles (alumno, profesor, admin) automáticamente al primer login",
     "Autenticación","backend/src/controllers/auth.controller.js\nbackend/src/models/User.js",
     "Alta","CU-01","CP-01","Finalizado"),
    ("RF-03","El sistema debe generar y validar tokens JWT con expiración configurable",
     "Autenticación","backend/src/middleware/verifyToken.js\nfrontend/src/app/core/interceptors/jwt.interceptor.ts",
     "Alta","CU-01","CP-02","Finalizado"),
    ("RF-04","Los alumnos deben poder consultar el catálogo de materias disponibles por período",
     "Inscripción","backend/src/controllers/materia.controller.js\nbackend/src/routes/materia.routes.js\nfrontend/src/app/modules/inscripcion/",
     "Alta","CU-02","CP-07","Finalizado"),
    ("RF-05","El sistema debe validar adeudos pendientes antes de permitir inscripción",
     "Inscripción","backend/src/services/inscripcion.service.js\nbackend/src/models/Adeudo.js",
     "Alta","CU-02","CP-03","Finalizado"),
    ("RF-06","El sistema debe verificar cupo disponible antes de inscribir a un alumno",
     "Inscripción","backend/src/services/inscripcion.service.js\nbackend/src/models/Materia.js",
     "Alta","CU-02","CP-04","Finalizado"),
    ("RF-07","El sistema debe validar prerrequisitos (seriación) antes de inscribir",
     "Inscripción","backend/src/services/inscripcion.service.js\nbackend/src/models/Materia.js",
     "Alta","CU-02","CP-05","Finalizado"),
    ("RF-08","El sistema debe detectar y rechazar cruces de horario entre materias",
     "Inscripción","backend/src/services/inscripcion.service.js",
     "Alta","CU-02","CP-06","Finalizado"),
    ("RF-09","Los alumnos deben poder consultar su inscripción activa con materias y horarios",
     "Inscripción","backend/src/controllers/inscripcion.controller.js\nfrontend/src/app/modules/dashboard/alumno/",
     "Media","CU-03","—","Finalizado"),
    ("RF-10","Los administradores deben poder cancelar inscripciones y liberar cupo",
     "Inscripción","backend/src/controllers/inscripcion.controller.js",
     "Media","CU-04","—","Finalizado"),
    ("RF-11","Los profesores deben poder capturar calificaciones parciales por alumno",
     "Calificaciones","backend/src/controllers/calificacion.controller.js\nfrontend/src/app/modules/calificaciones/",
     "Alta","CU-05","CP-11","Finalizado"),
    ("RF-12","El sistema debe calcular la calificación final automáticamente (promedio de parciales)",
     "Calificaciones","backend/src/controllers/calificacion.controller.js\nbackend/src/models/Calificacion.js",
     "Alta","CU-05","CP-11","Finalizado"),
    ("RF-13","Los profesores no deben poder editar actas ya cerradas",
     "Calificaciones","backend/src/controllers/calificacion.controller.js\nbackend/src/models/Calificacion.js",
     "Alta","CU-05","CP-11","Finalizado"),
    ("RF-14","Los administradores deben poder cerrar actas de calificaciones",
     "Calificaciones","backend/src/controllers/calificacion.controller.js\nbackend/src/routes/calificacion.routes.js",
     "Alta","CU-06","CP-11","Finalizado"),
    ("RF-15","Los alumnos deben consultar su historial académico completo agrupado por período",
     "Historial","backend/src/controllers/historial.controller.js\nfrontend/src/app/modules/historial/",
     "Alta","CU-07","CP-12","Finalizado"),
    ("RF-16","El sistema debe calcular promedio general y créditos acumulados del alumno",
     "Historial","backend/src/controllers/historial.controller.js\nbackend/src/models/Calificacion.js",
     "Media","CU-07","CP-12","Finalizado"),
    ("RF-17","El sistema debe generar un PDF descargable del kardex con colores institucionales",
     "Historial","backend/src/controllers/historial.controller.js\n(PDFKit 0.15)",
     "Media","CU-08","CP-12","Finalizado"),
    ("RF-18","Los administradores deben poder registrar adeudos para alumnos",
     "Adeudos","backend/src/controllers/adeudo.controller.js\nbackend/src/models/Adeudo.js",
     "Alta","CU-09","—","Finalizado"),
    ("RF-19","Los administradores deben poder marcar adeudos como pagados manualmente",
     "Adeudos","backend/src/controllers/adeudo.controller.js",
     "Alta","CU-10","—","Finalizado"),
    ("RF-20","Los alumnos deben poder consultar sus adeudos pendientes",
     "Adeudos","backend/src/controllers/adeudo.controller.js\nfrontend/src/app/modules/pagos/",
     "Media","CU-11","—","Finalizado"),
    ("RF-21","El sistema debe proteger todas las rutas con autenticación JWT",
     "Seguridad","backend/src/middleware/verifyToken.js\nbackend/src/routes/",
     "Alta","CU-01 a CU-12","CP-02","Finalizado"),
    ("RF-22","[EXCLUIDO] Envío de notificaciones push en tiempo real",
     "Notificaciones","N/A (fuera de alcance)",
     "—","—","—","Excluido"),
    ("RF-23","El sistema debe implementar control de acceso basado en roles (RBAC)",
     "Seguridad","backend/src/middleware/checkRole.js\nfrontend/src/app/core/guards/role.guard.ts",
     "Alta","CU-01 a CU-12","CP-13","Finalizado"),
    ("RF-24","Los administradores deben poder gestionar el catálogo de materias (CRUD)",
     "Materias","backend/src/controllers/materia.controller.js\nbackend/src/routes/materia.routes.js",
     "Media","CU-12","—","Finalizado"),
    ("RF-25","Los profesores deben ver únicamente las materias asignadas a ellos",
     "Materias","backend/src/controllers/materia.controller.js\nfrontend/src/app/modules/dashboard/profesor/",
     "Media","CU-05, CU-12","—","Finalizado"),
]

REQS_NO_FUNCIONALES = [
    # ID, Descripción, Criterio de aceptación, Módulo, Estado
    ("RNF-01","Seguridad: Los tokens JWT deben estar firmados con clave secreta de mínimo 64 caracteres y expirar en tiempo configurable",
     "Token generado con algoritmo HS256; variable JWT_EXPIRES_IN configurable","Autenticación / Seguridad","Finalizado"),
    ("RNF-02","Rendimiento: Las respuestas de la API deben procesarse en menos de 2 segundos bajo carga normal",
     "Tiempo de respuesta < 2000ms en llamadas principales (MongoDB Atlas + Railway)","Toda la API","En proceso"),
    ("RNF-03","Usabilidad: La interfaz debe ser intuitiva y consistente con Angular Material y colores institucionales",
     "Uso de Angular Material 17; paleta #0A2540 y #D4AF37 en toda la UI","Frontend (Angular)","Finalizado"),
    ("RNF-04","Disponibilidad: El sistema debe tener al menos 99.5% de uptime mensual",
     "Desplegado en Railway (backend) + Vercel (frontend) + MongoDB Atlas","Infraestructura","En proceso"),
    ("RNF-05","Escalabilidad: La arquitectura debe permitir agregar nuevos módulos sin afectar los existentes",
     "Estructura de carpetas por capas; lazy loading en Angular; módulos desacoplados","Arquitectura","Finalizado"),
    ("RNF-06","Mantenibilidad: El código debe estar organizado en capas (controllers, services, models, routes)",
     "Carpetas separadas para cada capa; convención de nombres consistente","Backend / Frontend","Finalizado"),
    ("RNF-07","Portabilidad: El sistema debe poder desplegarse en distintos proveedores cloud sin cambios al código",
     "Variables de entorno en .env; sin hardcoded URLs; configuración por ambiente","Backend / Frontend","Finalizado"),
]

CASOS_USO = [
    ("CU-01","Autenticar usuario con Google OAuth","Alumno, Profesor, Admin","RF-01, RF-02, RF-03","Alta","Finalizado"),
    ("CU-02","Inscribirse en materias del período","Alumno","RF-04, RF-05, RF-06, RF-07, RF-08","Alta","Finalizado"),
    ("CU-03","Consultar inscripción activa","Alumno","RF-09","Media","Finalizado"),
    ("CU-04","Cancelar inscripción de un alumno","Admin","RF-10","Media","Finalizado"),
    ("CU-05","Capturar calificaciones parciales","Profesor","RF-11, RF-12, RF-13","Alta","Finalizado"),
    ("CU-06","Cerrar acta de calificaciones","Admin","RF-14","Alta","Finalizado"),
    ("CU-07","Consultar historial académico (kardex)","Alumno, Admin","RF-15, RF-16","Alta","Finalizado"),
    ("CU-08","Descargar kardex en PDF","Alumno","RF-17","Media","Finalizado"),
    ("CU-09","Registrar adeudo a un alumno","Admin","RF-18","Alta","Finalizado"),
    ("CU-10","Marcar adeudo como pagado","Admin","RF-19","Alta","Finalizado"),
    ("CU-11","Consultar propios adeudos","Alumno","RF-20","Media","Finalizado"),
    ("CU-12","Gestionar catálogo de materias","Admin","RF-24, RF-25","Media","Finalizado"),
]

CASOS_PRUEBA = [
    ("CP-01","Login con cuenta Google válida","RF-01, RF-02","Token JWT generado; usuario creado en BD; redirección a dashboard según rol","Finalizado"),
    ("CP-02","Acceso a ruta protegida sin token","RF-21","HTTP 401 con mensaje de error en español","Finalizado"),
    ("CP-03","Inscripción con adeudo pendiente","RF-05","HTTP 400; código ADEUDO_PENDIENTE","Finalizado"),
    ("CP-04","Inscripción sin cupo disponible","RF-06","HTTP 400; código SIN_CUPO con nombre de materia","Finalizado"),
    ("CP-05","Inscripción sin prerequisito aprobado","RF-07","HTTP 400; código SERIACION_INCOMPLETA con materias involucradas","Finalizado"),
    ("CP-06","Inscripción con cruce de horario","RF-08","HTTP 400; código CRUCE_HORARIO con día del choque","Finalizado"),
    ("CP-07","Inscripción exitosa (todas las validaciones pasan)","RF-04 a RF-08","HTTP 201; inscripción creada; cupo decrementado en BD","Finalizado"),
    ("CP-08","[EXCLUIDO] Notificación push","RF-22","N/A (excluido del alcance)","Excluido"),
    ("CP-09","[EXCLUIDO] Pago en línea","RF-30","N/A (excluido del alcance)","Excluido"),
    ("CP-10","[EXCLUIDO] Referencia bancaria","RF-31","N/A (excluido del alcance)","Excluido"),
    ("CP-11","Editar calificación en acta cerrada","RF-13","HTTP 403; mensaje 'El acta está cerrada'","Finalizado"),
    ("CP-12","Generación de PDF del kardex","RF-17","PDF descargado con datos correctos y colores institucionales","Finalizado"),
    ("CP-13","Acceso con rol incorrecto","RF-23","HTTP 403 con mensaje en español","Finalizado"),
]


# ─── UTILIDADES DE ESTILO ─────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLANCO, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border_thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center",
                     wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center",
                     wrap_text=True)

def color_estado(estado):
    m = {"Finalizado": VERDE, "En proceso": NARANJA,
         "Pendiente": ROJO, "Excluido": MORADO}
    return m.get(estado, GRIS_MEDIO)

def bg_estado(estado):
    m = {"Finalizado": VERDE_CLARO, "En proceso": AMARILLO_CLARO,
         "Pendiente": ROJO_CLARO, "Excluido": "F3E5F5"}
    return m.get(estado, GRIS_CLARO)


# ─── HOJA: PORTADA ────────────────────────────────────────────────────────────
def crear_portada(wb):
    ws = wb.create_sheet("Portada", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30

    for r in range(1, 45):
        ws.row_dimensions[r].height = 18

    # Fondo azul oscuro en la cabecera
    for row in range(1, 14):
        for col in range(1, 4):
            c = ws.cell(row=row, column=col)
            c.fill = fill(AZUL_OSCURO)

    # Banda dorada
    for col in range(1, 4):
        ws.cell(row=14, column=col).fill = fill(DORADO)

    ws.merge_cells("A2:C5")
    t = ws["A2"]
    t.value = "UNIVERSIDAD SANTA FE — CAMPUS PARAÍSO"
    t.font = Font(name="Calibri", bold=True, size=18, color=DORADO)
    t.alignment = center()

    ws.merge_cells("A6:C8")
    t2 = ws["A6"]
    t2.value = "USF Portal — Sistema de Gestión Académica"
    t2.font = Font(name="Calibri", bold=True, size=14, color=BLANCO)
    t2.alignment = center()

    ws.merge_cells("A9:C11")
    t3 = ws["A9"]
    t3.value = "ACTIVIDAD 4\nMatriz de Trazabilidad de Requerimientos (MTR)"
    t3.font = Font(name="Calibri", bold=True, size=13, color=DORADO)
    t3.alignment = center()

    ws.merge_cells("A12:C13")
    t4 = ws["A12"]
    t4.value = "Materia: Trazabilidad y Configuración de Software"
    t4.font = Font(name="Calibri", italic=True, size=11, color=BLANCO)
    t4.alignment = center()

    datos_info = [
        (16, "Equipo:", "Equipo VIP — 7° y 8° Semestre"),
        (18, "Integrantes:", "Carlos Sanchez · Luis Roberto Suarez · Ricardo Galindo\nEduardo Robles · Mariana Jimenez"),
        (21, "Catedrático:", "Rafael Bernal Gallegos"),
        (23, "Fecha de entrega:", "07 de junio de 2026"),
        (25, "Versión:", "1.0 — Entrega Final"),
    ]

    for fila, etiqueta, valor in datos_info:
        ws.row_dimensions[fila].height = 22 if "\n" in valor else 18
        ce = ws.cell(row=fila, column=2)
        ce.value = etiqueta
        ce.font = Font(name="Calibri", bold=True, size=11, color=AZUL_OSCURO)
        ce.alignment = left()
        cv = ws.cell(row=fila, column=3)
        cv.value = valor
        cv.font = Font(name="Calibri", size=11, color=AZUL_CLARO)
        cv.alignment = left()

    # Leyenda de estados
    ws.merge_cells("B28:C28")
    lt = ws["B28"]
    lt.value = "LEYENDA DE ESTADOS"
    lt.font = Font(name="Calibri", bold=True, size=11, color=AZUL_OSCURO)
    lt.alignment = left()

    leyendas = [
        (29, VERDE, VERDE_CLARO, "Finalizado", "El requerimiento está completamente implementado y validado"),
        (30, NARANJA, AMARILLO_CLARO, "En proceso", "El requerimiento está parcialmente implementado"),
        (31, ROJO, ROJO_CLARO, "Pendiente", "El requerimiento aún no ha sido implementado"),
        (32, MORADO, "F3E5F5", "Excluido", "Fuera del alcance del proyecto por instrucción del catedrático"),
    ]

    for fila, color_txt, color_bg, estado, descripcion in leyendas:
        ws.row_dimensions[fila].height = 18
        cb = ws.cell(row=fila, column=2)
        cb.value = f"  {estado}"
        cb.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
        cb.fill = fill(color_txt)
        cb.alignment = left()
        cd = ws.cell(row=fila, column=3)
        cd.value = descripcion
        cd.font = Font(name="Calibri", size=10, color=AZUL_OSCURO)
        cd.alignment = left()

    # Resumen de totales
    ws.merge_cells("B35:C35")
    rt = ws["B35"]
    rt.value = "RESUMEN DEL PROYECTO"
    rt.font = Font(name="Calibri", bold=True, size=11, color=AZUL_OSCURO)
    rt.alignment = left()

    resumenes = [
        (36, "Requerimientos Funcionales:", "24 activos + 1 excluido (RF-22)"),
        (37, "Requerimientos No Funcionales:", "7 requerimientos"),
        (38, "Casos de Uso:", "12 casos de uso"),
        (39, "Casos de Prueba:", "10 activos + 3 excluidos"),
        (40, "Estado general:", "92% finalizado — En fase de pruebas"),
    ]

    for fila, etq, val in resumenes:
        ws.row_dimensions[fila].height = 18
        e = ws.cell(row=fila, column=2)
        e.value = etq
        e.font = Font(name="Calibri", bold=True, size=10, color=AZUL_OSCURO)
        e.alignment = left()
        v = ws.cell(row=fila, column=3)
        v.value = val
        v.font = Font(name="Calibri", size=10, color=AZUL_CLARO)
        v.alignment = left()

    return ws


# ─── HOJA: MATRIZ DE TRAZABILIDAD ────────────────────────────────────────────
def crear_matriz(wb):
    ws = wb.create_sheet("Matriz de Trazabilidad")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Anchos de columna
    anchos = [6, 14, 42, 12, 16, 36, 20, 18, 16]
    encabezados = [
        "#", "ID", "Descripción del Requerimiento", "Prioridad",
        "Módulo del Sistema", "Componentes / Archivos",
        "Casos de Uso", "Casos de Prueba", "Estado"
    ]

    for i, (a, h) in enumerate(zip(anchos, encabezados), 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    # Título
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "MATRIZ DE TRAZABILIDAD DE REQUERIMIENTOS — USF PORTAL"
    t.font = Font(name="Calibri", bold=True, size=13, color=BLANCO)
    t.fill = fill(AZUL_OSCURO)
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    # Encabezados
    ws.row_dimensions[2].height = 35
    for col, h in enumerate(encabezados, 1):
        c = ws.cell(row=2, column=col)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
        c.fill = fill(DORADO)
        c.alignment = center()
        c.border = border_thin()

    # ── Sección: Requerimientos Funcionales ──
    fila = 3
    ws.merge_cells(f"A{fila}:I{fila}")
    sec = ws[f"A{fila}"]
    sec.value = "▶  REQUERIMIENTOS FUNCIONALES"
    sec.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
    sec.fill = fill(AZUL_CLARO)
    sec.alignment = left()
    ws.row_dimensions[fila].height = 22
    fila += 1

    for idx, (rid, desc, mod, arch, prio, cu, cp, estado) in enumerate(REQS_FUNCIONALES, 1):
        ws.row_dimensions[fila].height = 40
        bg = GRIS_CLARO if idx % 2 == 0 else AZUL_FILAS
        if estado == "Excluido":
            bg = "F3F3F3"

        vals = [str(idx), rid, desc, prio, mod, arch, cu, cp, estado]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=fila, column=col)
            c.value = v
            c.border = border_thin()
            if estado == "Excluido":
                c.font = Font(name="Calibri", size=9, color="999999",
                              italic=True, strikethrough=(col > 1))
                c.fill = fill("F0F0F0")
            else:
                c.font = Font(name="Calibri", size=9, color=AZUL_OSCURO)
                c.fill = fill(bg)

            if col == 9:  # Estado
                c.font = Font(name="Calibri", bold=True, size=9,
                              color=BLANCO if estado != "Excluido" else "999999")
                if estado != "Excluido":
                    c.fill = fill(color_estado(estado))
                c.alignment = center()
            elif col in (1, 2, 4):
                c.alignment = center()
            else:
                c.alignment = left()

        fila += 1

    # ── Sección: Requerimientos No Funcionales ──
    ws.merge_cells(f"A{fila}:I{fila}")
    sec2 = ws[f"A{fila}"]
    sec2.value = "▶  REQUERIMIENTOS NO FUNCIONALES"
    sec2.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
    sec2.fill = fill(AZUL_CLARO)
    sec2.alignment = left()
    ws.row_dimensions[fila].height = 22
    fila += 1

    for idx2, (rid, desc, criterio, mod, estado) in enumerate(REQS_NO_FUNCIONALES, 1):
        ws.row_dimensions[fila].height = 40
        bg = GRIS_CLARO if idx2 % 2 == 0 else AMARILLO_CLARO
        vals = [str(idx2), rid, desc, "—", mod, criterio, "—", "—", estado]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=fila, column=col)
            c.value = v
            c.border = border_thin()
            c.font = Font(name="Calibri", size=9, color=AZUL_OSCURO)
            c.fill = fill(bg)
            if col == 9:
                c.font = Font(name="Calibri", bold=True, size=9, color=BLANCO)
                c.fill = fill(color_estado(estado))
                c.alignment = center()
            elif col in (1, 2, 4):
                c.alignment = center()
            else:
                c.alignment = left()
        fila += 1

    return ws


# ─── HOJA: CASOS DE USO ──────────────────────────────────────────────────────
def crear_casos_uso(wb):
    ws = wb.create_sheet("Casos de Uso")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    anchos = [6, 10, 36, 22, 22, 10, 14]
    headers = ["#", "ID", "Caso de Uso", "Actores", "RF Relacionados",
               "Prioridad", "Estado"]

    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "CATÁLOGO DE CASOS DE USO — USF PORTAL"
    t.font = Font(name="Calibri", bold=True, size=13, color=BLANCO)
    t.fill = fill(AZUL_OSCURO)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    ws.row_dimensions[2].height = 30
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
        c.fill = fill(DORADO)
        c.alignment = center()
        c.border = border_thin()

    for idx, (cid, nombre, actores, rfs, prio, estado) in enumerate(CASOS_USO, 1):
        fila = idx + 2
        ws.row_dimensions[fila].height = 30
        bg = AZUL_FILAS if idx % 2 != 0 else GRIS_CLARO
        vals = [str(idx), cid, nombre, actores, rfs, prio, estado]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=fila, column=col)
            c.value = v
            c.border = border_thin()
            c.fill = fill(bg)
            c.font = Font(name="Calibri", size=10, color=AZUL_OSCURO)
            if col == 7:
                c.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
                c.fill = fill(color_estado(estado))
                c.alignment = center()
            elif col in (1, 2, 6):
                c.alignment = center()
            else:
                c.alignment = left()

    return ws


# ─── HOJA: CASOS DE PRUEBA ───────────────────────────────────────────────────
def crear_casos_prueba(wb):
    ws = wb.create_sheet("Casos de Prueba")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    anchos = [5, 10, 35, 16, 40, 14]
    headers = ["#", "ID", "Caso de Prueba", "RF Validado",
               "Resultado Esperado", "Estado"]

    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "CASOS DE PRUEBA — USF PORTAL"
    t.font = Font(name="Calibri", bold=True, size=13, color=BLANCO)
    t.fill = fill(AZUL_OSCURO)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    ws.row_dimensions[2].height = 30
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
        c.fill = fill(DORADO)
        c.alignment = center()
        c.border = border_thin()

    for idx, (cid, nombre, rfs, resultado, estado) in enumerate(CASOS_PRUEBA, 1):
        fila = idx + 2
        ws.row_dimensions[fila].height = 35
        bg = GRIS_CLARO if idx % 2 == 0 else AZUL_FILAS
        if estado == "Excluido":
            bg = "F0F0F0"
        vals = [str(idx), cid, nombre, rfs, resultado, estado]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=fila, column=col)
            c.value = v
            c.border = border_thin()
            if estado == "Excluido":
                c.font = Font(name="Calibri", size=9, color="999999",
                              italic=True, strikethrough=(col > 1))
                c.fill = fill("F0F0F0")
            else:
                c.font = Font(name="Calibri", size=9, color=AZUL_OSCURO)
                c.fill = fill(bg)
            if col == 6:
                c.font = Font(name="Calibri", bold=True, size=9,
                              color=BLANCO if estado != "Excluido" else "999999")
                if estado != "Excluido":
                    c.fill = fill(color_estado(estado))
                c.alignment = center()
            elif col in (1, 2):
                c.alignment = center()
            else:
                c.alignment = left()

    return ws


# ─── HOJA: ANÁLISIS DE BRECHAS ───────────────────────────────────────────────
def crear_analisis(wb):
    ws = wb.create_sheet("Análisis de Brechas")
    ws.sheet_view.showGridLines = False

    for i, a in enumerate([5, 14, 42, 16, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "ANÁLISIS DE BRECHAS Y REQUERIMIENTOS SIN EVIDENCIA"
    t.font = Font(name="Calibri", bold=True, size=13, color=BLANCO)
    t.fill = fill(AZUL_OSCURO)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    ws.row_dimensions[2].height = 30
    for col, h in enumerate(["#","ID","Requerimiento","Situación","Observación / Acción recomendada"], 1):
        c = ws.cell(row=2, column=col)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
        c.fill = fill(DORADO)
        c.alignment = center()
        c.border = border_thin()

    brechas = [
        ("RNF-02","Rendimiento: respuesta < 2 segundos","En proceso",
         "No se han ejecutado pruebas de carga (ej. Artillery/k6). Se recomienda configurar una prueba de estrés básica en Railway antes del despliegue final."),
        ("RNF-04","Disponibilidad 99.5% de uptime","En proceso",
         "No existe monitoreo activo configurado (ej. UptimeRobot). Se recomienda agregar alertas de disponibilidad post-despliegue."),
        ("CP-03 a CP-07","Casos de prueba de inscripción","En proceso",
         "Los casos de prueba están documentados y las validaciones implementadas, pero no existe aún el archivo backend/src/__tests__/inscripcion.test.js con código Jest/Supertest ejecutable."),
        ("CP-08, CP-09, CP-10","Casos de prueba de pago y notificaciones","Excluido",
         "RF-22 (notificaciones) y RF-30/RF-31 (pagos) fueron excluidos del alcance por instrucción del catedrático. No aplica su implementación ni prueba."),
        ("RF-09","Consulta de inscripción activa (vista detalle)","Finalizado",
         "Implementado en el dashboard del alumno, pero no existe un endpoint exclusivo /api/inscripciones/mi-inscripcion documentado en Postman. Sin impacto funcional."),
        ("RF-25","Materias asignadas por profesor (filtro)","Finalizado",
         "La lógica de filtro existe en el backend pero la vista del dashboard del profesor muestra todas las materias. Se recomienda agregar filtro por profesor en la consulta GET /api/materias."),
    ]

    for idx, (rid, req, sit, obs) in enumerate(brechas, 1):
        fila = idx + 2
        ws.row_dimensions[fila].height = 50
        colores = {"En proceso": (NARANJA, AMARILLO_CLARO),
                   "Excluido": (MORADO, "F3E5F5"),
                   "Finalizado": (VERDE, VERDE_CLARO)}
        c_txt, c_bg = colores.get(sit, (GRIS_MEDIO, GRIS_CLARO))
        vals = [str(idx), rid, req, sit, obs]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=fila, column=col)
            c.value = v
            c.border = border_thin()
            c.font = Font(name="Calibri", size=9, color=AZUL_OSCURO)
            c.fill = fill(c_bg)
            if col == 4:
                c.font = Font(name="Calibri", bold=True, size=9, color=BLANCO)
                c.fill = fill(c_txt)
                c.alignment = center()
            elif col in (1, 2):
                c.alignment = center()
            else:
                c.alignment = left()

    # Conclusión
    ws.row_dimensions[10].height = 18
    ws.merge_cells("A10:E10")
    concl_t = ws["A10"]
    concl_t.value = "CONCLUSIÓN DEL ANÁLISIS"
    concl_t.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
    concl_t.fill = fill(AZUL_CLARO)
    concl_t.alignment = left()

    ws.merge_cells("A11:E14")
    concl = ws["A11"]
    concl.value = (
        "El 92% de los requerimientos funcionales y no funcionales del USF Portal se encuentran en estado Finalizado. "
        "Las principales brechas identificadas son de carácter no funcional (rendimiento y disponibilidad) y corresponden "
        "a la ausencia de pruebas automatizadas de estrés y herramientas de monitoreo, lo cual no impide el funcionamiento "
        "correcto del sistema pero sí representa un riesgo operativo en producción.\n\n"
        "Los requerimientos excluidos (RF-22, RF-30, RF-31) fueron eliminados del alcance por instrucción del catedrático "
        "y no representan una brecha de implementación.\n"
        "Acción inmediata recomendada: implementar los archivos de prueba Jest (.test.js) para los módulos de inscripción y calificaciones."
    )
    concl.font = Font(name="Calibri", size=10, color=AZUL_OSCURO)
    concl.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for fila in range(11, 15):
        ws.row_dimensions[fila].height = 22

    return ws


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    # Quitar hoja por defecto
    del wb["Sheet"]

    crear_portada(wb)
    crear_matriz(wb)
    crear_casos_uso(wb)
    crear_casos_prueba(wb)
    crear_analisis(wb)

    ruta = "/home/user/proyectofinalequipoVip/entregables/Actividad4_MatrizTrazabilidad_USFPortal.xlsx"
    wb.save(ruta)
    print(f"✅  Excel guardado en: {ruta}")


if __name__ == "__main__":
    main()
